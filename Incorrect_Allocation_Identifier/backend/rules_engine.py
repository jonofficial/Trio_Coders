"""
rules_engine.py — Business Rules Engine & Exception Report Generator

Evaluates all seven business rules (R1–R7) against merged RIS+SO data.
Each rule is implemented as a separate method for testability.
Generates an Excel exception report with full output schema compliance.
"""

import pandas as pd
import numpy as np
import logging
import os
from datetime import datetime, date

logger = logging.getLogger(__name__)


class RulesEngine:
    """Evaluates R1–R7 business rules and generates the exception report."""

    # Output schema column order (per PROJECT_CONTEXT-2 Section 6.2)
    OUTPUT_COLUMNS = [
        'Emplid', 'Employee Name', 'Project Id', 'Project Name',
        'Task Type', '% Allocation', 'Computed_Total_Allocation', 'Allocation % Status',
        'Rule ID', 'Rule Name', 'Severity', 'Exception Description',
        'Recommended Action',
        'Project Type', 'Project Role', 'SO ID', 'SO Category',
        'Mode of Hire', 'Employee Status', 'Grade',
        'Location Category', 'Onsite Status',
        'Current or Future Bench', 'Bench Flag',
        'Resource End Date', 'Days Past End Date',
        'Resource Based SL', 'Resource Based SSL', 'SO Resource Base SSL',
        'RM', 'Supervisor Name', 'Supervisor Email', 'Employee Email',
        'Run Date',
    ]

    def __init__(self, merged_df: pd.DataFrame = None, data_path: str = None,
                 contractor_cfg_path: str = None, role_cfg_path: str = None,
                 contractor_codes: frozenset = None, support_sales_roles: frozenset = None):
        """
        Initialize with either a DataFrame directly or a CSV path (legacy).
        Config values can be passed directly or loaded from CSV files (legacy).
        """
        self._merged_df = merged_df
        self._data_path = data_path
        self._contractor_cfg_path = contractor_cfg_path
        self._role_cfg_path = role_cfg_path
        self._contractor_codes = contractor_codes or frozenset()
        self._support_sales_roles = support_sales_roles or frozenset()

    def load_data(self):
        """Legacy method: load data and configs from file paths."""
        if self._data_path:
            logger.info(f"Loading prepared data from {self._data_path}")
            self._merged_df = pd.read_csv(self._data_path)

            if 'Resource End Date' in self._merged_df.columns:
                self._merged_df['Resource End Date'] = pd.to_datetime(
                    self._merged_df['Resource End Date'], errors='coerce'
                )

        if self._contractor_cfg_path:
            self._contractor_codes = frozenset(
                pd.read_csv(self._contractor_cfg_path).iloc[:, 0].dropna().astype(str).str.strip().str.lower()
            )
        if self._role_cfg_path:
            self._support_sales_roles = frozenset(
                pd.read_csv(self._role_cfg_path).iloc[:, 0].dropna().astype(str).str.strip().str.lower()
            )

    def evaluate_rules(self) -> pd.DataFrame:
        """Evaluate all 7 business rules. Returns a single exceptions DataFrame."""
        if self._merged_df is None or self._merged_df.empty:
            logger.warning("No data to evaluate. Returning empty exceptions.")
            return pd.DataFrame(columns=self.OUTPUT_COLUMNS)

        df = self._merged_df.copy()
        logger.info(f"Evaluating rules on {len(df)} records...")

        # Resolve column names (may have _RIS suffix from merge)
        self._task_col = self._resolve_col(df, 'Task Type_RIS', 'Task Type')
        self._alloc_col = self._resolve_col(df, '% Allocation_RIS', '% Allocation')
        self._role_col = self._resolve_col(df, 'Project Role_RIS', 'Project Role')
        self._rssl_col = self._resolve_col(df, 'Resource Based SSL_RIS', 'Resource Based SSL', 'Resource Based SSL ')

        # Ensure _bench_flag exists (fallback if data came from legacy CSV without it)
        if '_bench_flag' not in df.columns:
            df['_bench_flag'] = self._compute_bench_fallback(df)

        if '_end_date_parseable' not in df.columns:
            df['_end_date_parseable'] = df['Resource End Date'].notna() if 'Resource End Date' in df.columns else False

        # Evaluate each rule
        all_exceptions = []
        all_exceptions.extend(self._rule_r1(df))
        all_exceptions.extend(self._rule_r2(df))
        all_exceptions.extend(self._rule_r3(df))
        all_exceptions.extend(self._rule_r4(df))
        all_exceptions.extend(self._rule_r5(df))
        all_exceptions.extend(self._rule_r6(df))
        all_exceptions.extend(self._rule_r7(df))

        exceptions_df = pd.DataFrame(all_exceptions)

        # Ensure output schema compliance
        if not exceptions_df.empty:
            for col in self.OUTPUT_COLUMNS:
                if col not in exceptions_df.columns:
                    exceptions_df[col] = 'N/A'
            exceptions_df = exceptions_df[self.OUTPUT_COLUMNS]
            exceptions_df.fillna('N/A', inplace=True)

        logger.info(f"Rule evaluation complete. Total exceptions: {len(exceptions_df)}")
        return exceptions_df

    # ==================== R1: Unassigned Task ====================

    def _rule_r1(self, df: pd.DataFrame) -> list:
        """R1: Flag records where Task Type is 'UN'. Sub-classify as pure vs mixed."""
        exceptions = []
        mask = df[self._task_col].astype(str).str.strip().str.upper() == 'UN'
        un_records = df[mask]

        if un_records.empty:
            logger.info("R1: No unassigned task violations found.")
            return exceptions

        # Determine which PSIDs are pure-unassigned vs mixed
        for psid in un_records['Emplid'].unique():
            psid_records = df[df['Emplid'] == psid]
            psid_un_records = un_records[un_records['Emplid'] == psid]
            all_un = psid_records[self._task_col].astype(str).str.strip().str.upper().eq('UN').all()

            if all_un:
                sub_class = "Fully unassigned — no valid project context"
            else:
                sub_class = "Unassigned record — mixed assignment"

            for _, row in psid_un_records.iterrows():
                exc = self._format_exception(
                    row, "R1", "Unassigned Task", "High",
                    f"Task Type is 'UN'. {sub_class}",
                    "Assign a valid Task Type or close the unassigned record"
                )
                exceptions.append(exc)

        logger.info(f"R1: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R2: Incorrect Task Tag ====================

    def _rule_r2(self, df: pd.DataFrame) -> list:
        """R2: Flag records where role is Support/Sales but tagged NE instead of NI."""
        exceptions = []

        if self._role_col not in df.columns:
            logger.warning("R2: Project Role column not found. Skipping.")
            return exceptions

        def is_support_sales(val):
            s = str(val).strip().lower()
            return any(role in s for role in self._support_sales_roles)

        is_ne = df[self._task_col].astype(str).str.strip().str.upper() == 'NE'
        is_role = df[self._role_col].apply(is_support_sales)
        mask = is_ne & is_role

        for _, row in df[mask].iterrows():
            exc = self._format_exception(
                row, "R2", "Incorrect Task Tag", "Medium",
                f"Support/Sales role '{row.get(self._role_col, 'N/A')}' incorrectly tagged as NE; should be NI",
                "Change Task Type from NE to NI"
            )
            exceptions.append(exc)

        logger.info(f"R2: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R3: Allocation ≠ 100% ====================

    def _rule_r3(self, df: pd.DataFrame) -> list:
        """R3: Flag persons whose total allocation ≠ 100%. One row per PSID."""
        exceptions = []

        # Use pre-computed Computed_Total_Allocation if available, else compute
        if 'Computed_Total_Allocation' in df.columns:
            person_alloc = df.drop_duplicates(subset=['Emplid'])[['Emplid', 'Computed_Total_Allocation']].copy()
        else:
            person_alloc = df.groupby('Emplid')[self._alloc_col].sum().reset_index()
            person_alloc.rename(columns={self._alloc_col: 'Computed_Total_Allocation'}, inplace=True)

        # Round to 2 decimal places for floating-point tolerance
        person_alloc['Computed_Total_Allocation'] = person_alloc['Computed_Total_Allocation'].round(2)

        # Flag under and over allocation
        violations = person_alloc[person_alloc['Computed_Total_Allocation'] != 100.0]

        for _, person in violations.iterrows():
            psid = person['Emplid']
            total = person['Computed_Total_Allocation']

            if total < 100.0:
                alloc_type = "Under-Allocated"
            else:
                alloc_type = "Over-Allocated"

            # Get a representative row for this PSID (first record)
            rep_row = df[df['Emplid'] == psid].iloc[0]

            exc = self._format_exception(
                rep_row, "R3", "Allocation != 100%", "Critical",
                f"{alloc_type}: Total allocation is {total}% (expected 100%)",
                "Adjust allocation percentages to sum to exactly 100%"
            )
            exc['Computed_Total_Allocation'] = total
            exc['Allocation % Status'] = alloc_type
            exceptions.append(exc)

        logger.info(f"R3: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R4: Contractor on Bench ====================

    def _rule_r4(self, df: pd.DataFrame) -> list:
        """R4: Flag contractors who are on bench."""
        exceptions = []

        # Determine contractor status from Employee Status or Mode of Hire
        status_col = self._resolve_col(df, 'Employee Status', 'Mode of Hire', 'Personnel Status')

        if status_col not in df.columns:
            logger.warning("R4: No employee status/hiring category column found. Skipping.")
            return exceptions

        def is_contractor(val):
            s = str(val).strip().lower()
            if s in ('nan', '', 'n/a'):
                return False
            return s in self._contractor_codes

        is_contr = df[status_col].apply(is_contractor)
        mask = is_contr & df['_bench_flag']

        for _, row in df[mask].iterrows():
            exc = self._format_exception(
                row, "R4", "Contractor on Bench", "High",
                f"Contractor (category: '{row.get(status_col, 'N/A')}') is on bench",
                "Review contractor bench status; initiate offboarding or re-assignment"
            )
            exceptions.append(exc)

        logger.info(f"R4: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R5: Onsite on Bench ====================

    def _rule_r5(self, df: pd.DataFrame) -> list:
        """R5: Flag employees who are onsite while on bench."""
        exceptions = []

        loc_col = self._resolve_col(df, 'Location Category ', 'Location Category', 'Onsite Status')
        if loc_col not in df.columns:
            logger.warning("R5: No location column found. Skipping.")
            return exceptions

        def is_onsite(val):
            s = str(val).strip().lower()
            return 'onsite' in s

        mask = df[loc_col].apply(is_onsite) & df['_bench_flag']

        for _, row in df[mask].iterrows():
            exc = self._format_exception(
                row, "R5", "Onsite on Bench", "High",
                f"Employee is Onsite ('{row.get(loc_col, 'N/A')}') while on bench",
                "Recall employee from client site or update bench status"
            )
            exceptions.append(exc)

        logger.info(f"R5: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R6: Stale Long Leave ====================

    def _rule_r6(self, df: pd.DataFrame) -> list:
        """R6: Flag long-leave records where end date has passed and employee is on bench."""
        exceptions = []

        # Use system date for evaluation per spec
        today = pd.Timestamp(datetime.now().date())

        is_ll = df[self._task_col].astype(str).str.strip().str.upper() == 'LL'
        is_bench = df['_bench_flag']
        is_parseable = df['_end_date_parseable']
        is_past = df['Resource End Date'] < today

        mask = is_ll & is_bench & is_parseable & is_past

        for _, row in df[mask].iterrows():
            end_date = row['Resource End Date']
            days_past = (today - end_date).days if pd.notna(end_date) else 0

            exc = self._format_exception(
                row, "R6", "Stale Long Leave", "Medium",
                f"Long leave end date ({end_date.strftime('%Y-%m-%d') if pd.notna(end_date) else 'Unknown'}) "
                f"is {days_past} day(s) past. Employee still on bench.",
                "Close or extend the long-leave record; update bench status"
            )
            exc['Days Past End Date'] = days_past
            exceptions.append(exc)

        logger.info(f"R6: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== R7: Missing RSSL ====================

    def _rule_r7(self, df: pd.DataFrame) -> list:
        """R7: Flag records where Resource Based SSL is missing, null, or invalid."""
        exceptions = []

        if self._rssl_col not in df.columns:
            logger.warning(f"R7: RSSL column '{self._rssl_col}' not found. Skipping.")
            return exceptions

        def is_missing_rssl(val):
            if pd.isna(val):
                return True
            s = str(val).strip().lower()
            return s in ('', 'nan', 'null', 'invalid', 'n/a', 'none')

        mask = df[self._rssl_col].apply(is_missing_rssl)

        # SO RSSL column for mismatch classification
        so_rssl_col = self._resolve_col(df, 'Resource Base SSL', 'Resource Base SSL_SO', 'Resource Based SSL_SO')

        for _, row in df[mask].iterrows():
            # Sub-classify: RIS-SO Mismatch vs Missing in RIS
            sub_class = "Missing in RIS"
            if so_rssl_col in df.columns:
                so_val = row.get(so_rssl_col, None)
                if pd.notna(so_val) and str(so_val).strip().lower() not in ('', 'nan', 'n/a', 'none', 'null'):
                    sub_class = "RIS–SO Mismatch"

            exc = self._format_exception(
                row, "R7", "Missing RSSL", "Medium",
                f"Resource Based SSL is missing or invalid. Classification: {sub_class}",
                "Populate Resource Based SSL in RIS; reconcile with SO where mismatch exists"
            )
            exc['SO Resource Base SSL'] = row.get(so_rssl_col, 'N/A') if so_rssl_col in df.columns else 'N/A'
            exceptions.append(exc)

        logger.info(f"R7: {len(exceptions)} exception(s) found.")
        return exceptions

    # ==================== Helpers ====================

    def _resolve_col(self, df: pd.DataFrame, *candidates: str) -> str:
        """Return the first column name that exists in the DataFrame."""
        for col in candidates:
            if col in df.columns:
                return col
        return candidates[0]  # Return first candidate as default

    def _compute_bench_fallback(self, df: pd.DataFrame) -> pd.Series:
        """Fallback bench detection for legacy CSV data without _bench_flag."""
        def _check_bench(row):
            for col in ['Current or Future Bench', 'Current or Future Bench ']:
                if col in row.index and pd.notna(row[col]) and str(row[col]).strip() != '':
                    s = str(row[col]).strip().lower()
                    if s in ('yes', 'y', 'true', '1') or 'bench' in s:
                        return True
                    return False
            for col in ['Bench Allocation Category', 'Bench Allocation Category ']:
                if col in row.index and pd.notna(row[col]) and str(row[col]).strip() != '':
                    return 'bench' in str(row[col]).strip().lower()
            for col in ['Project Type', 'Project Category ', 'Project Category']:
                if col in row.index and pd.notna(row[col]):
                    if 'bench' in str(row[col]).strip().lower():
                        return True
            return False
        return df.apply(_check_bench, axis=1)

    def _format_exception(self, row, rule_id: str, rule_name: str, severity: str,
                          description: str, recommended_action: str) -> dict:
        """Format a single exception row with full output schema. Never leave cells blank."""

        def safe_get(key, *fallbacks):
            """Safely extract a value from the row, trying fallback keys."""
            for k in [key] + list(fallbacks):
                if k in row.index:
                    val = row[k]
                    if pd.notna(val) and str(val).strip() not in ('', 'nan'):
                        return str(val).strip()
            return 'N/A'

        return {
            'Emplid': safe_get('Emplid'),
            'Employee Name': safe_get('Employee Name'),
            'Project Id': safe_get('Project Id', 'Project Id_RIS'),
            'Project Name': safe_get('Project Name', 'Project Name_RIS'),
            'Task Type': safe_get(self._task_col, 'Task Type'),
            '% Allocation': safe_get(self._alloc_col, '% Allocation'),
            'Computed_Total_Allocation': safe_get('Computed_Total_Allocation'),
            'Allocation % Status': safe_get('Allocation % Status'),
            'Rule ID': rule_id,
            'Rule Name': rule_name,
            'Severity': severity,
            'Exception Description': description,
            'Recommended Action': recommended_action,
            'Project Type': safe_get('Project Type'),
            'Project Role': safe_get(self._role_col, 'Project Role'),
            'SO ID': safe_get('SO ID', 'SO Id'),
            'SO Category': safe_get('SO Category'),
            'Mode of Hire': safe_get('Mode of Hire'),
            'Employee Status': safe_get('Employee Status'),
            'Grade': safe_get('Grade'),
            'Location Category': safe_get('Location Category ', 'Location Category'),
            'Onsite Status': safe_get('Onsite Status'),
            'Current or Future Bench': safe_get('Current or Future Bench'),
            'Bench Flag': str(row.get('_bench_flag', 'N/A')),
            'Resource End Date': safe_get('Resource End Date'),
            'Days Past End Date': 'N/A',
            'Resource Based SL': safe_get('Resource Based SL', 'Resource Based SL_RIS'),
            'Resource Based SSL': safe_get(self._rssl_col, 'Resource Based SSL'),
            'SO Resource Base SSL': safe_get('Resource Base SSL', 'Resource Base SSL_SO'),
            'RM': safe_get('RM'),
            'Supervisor Name': safe_get('Supervisor Name'),
            'Supervisor Email': safe_get('Supervisor Email ID'),
            'Employee Email': safe_get('Employee Email ID'),
            'Run Date': datetime.now().isoformat(),
        }

    def generate_report(self, exceptions_df: pd.DataFrame, output_dir: str = None) -> str:
        """Generate the Excel exception report with formatting."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"IAI_ExceptionReport_{timestamp}.xlsx"

        if output_dir is None:
            output_dir = os.path.dirname(self._data_path) if self._data_path else "."
        output_path = os.path.join(output_dir, output_filename)

        logger.info(f"Generating Exception Report: {output_filename}")

        # Fill blanks with 'N/A'
        exceptions_df = exceptions_df.fillna('N/A')

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            exceptions_df.to_excel(writer, sheet_name='IAI_Exceptions', index=False)

            # Apply severity color-coding
            try:
                from openpyxl.styles import PatternFill, Font
                ws = writer.sheets['IAI_Exceptions']

                # Bold header row
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Color-code severity column
                severity_col_idx = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Severity':
                        severity_col_idx = idx
                        break

                if severity_col_idx:
                    color_map = {
                        'Critical': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
                        'High': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
                        'Medium': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
                    }
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=severity_col_idx)
                        if cell.value in color_map:
                            cell.fill = color_map[cell.value]

            except Exception as e:
                logger.warning(f"Could not apply Excel formatting: {e}")

        logger.info(f"Report saved at {output_path}")
        return output_path


if __name__ == "__main__":
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(os.path.dirname(CURRENT_DIR), "Data")

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    engine = RulesEngine(
        data_path=os.path.join(DATA_DIR, "Merged_Prepared_Data.csv"),
        contractor_cfg_path=os.path.join(DATA_DIR, "contractor_config.csv"),
        role_cfg_path=os.path.join(DATA_DIR, "role_mapping_config.csv")
    )
    engine.load_data()
    exceptions_df = engine.evaluate_rules()
    engine.generate_report(exceptions_df, output_dir=DATA_DIR)
